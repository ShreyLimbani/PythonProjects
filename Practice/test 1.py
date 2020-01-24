import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('Consolidated_BTech_Project_Review_March 2017-18.xlsx')
sheet = wb['Batch-1']
wb2 = xl.load_workbook('./Project work - 2014/2014 Project - IT/2014-(NW).xlsx')
sheet2 = wb2['CIA-1-Component 1']

#for row in sheet2.iter_rows(min_row=1, max_col=3, max_row=2):
#    for cell in row:
#        print(cell)

for row in range(5, 23):
    found = False
    cell1 = sheet2.cell(row, 2)
    v = cell1.value
    print(v)
    for row2 in range(4,56):
        cell2 = sheet.cell(row2, 4)
        #print(cell2.value)
        if cell1.value == sheet.cell(row2, 4).value:
            found = True
            sheet.cell(row2, 8).value = '--NW--'
            break

    if found:
        continue

    sheet = wb['Batch_2']
    for row2 in range(4, 55):
        cell2 = sheet.cell(row2, 4)
        # print(cell2.value)
        if cell1.value == sheet.cell(row2, 4).value:
            found = True
            sheet.cell(row2, 8).value = '--NW--'
            break

    if found:
        continue

    sheet = wb['Batch-3']
    for row2 in range(4, 38):
        cell2 = sheet.cell(row2, 4)
        # print(cell2.value)
        if cell1.value == sheet.cell(row2, 4).value:
            found = True
            sheet.cell(row2, 8).value = '--NW--'
            break

wb.save('Consolidated_BTech_Project_Review_March 2017-18.xlsx')