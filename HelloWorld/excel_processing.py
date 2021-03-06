import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value * 0.5
    sheet.cell(row, 4).value = corrected_price

wb.save('transactions(2).xlsx')
